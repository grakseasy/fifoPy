#Define the covergroup name
from openpyxl import load_workbook
import re

wb = load_workbook('fifo_cg.xlsx')
ws = wb['Sheet1'] 

column_data = []
for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=15, values_only=True):
    for cell_value in row:
        if cell_value is not None:
            column_data.append(cell_value)

covergroup_name = column_data.pop(0)
column_data.pop(0) #This leaves me with only SPLIT_BINS columns

signals = []
for signal in column_data:
    string = signal.split(":",1)[1]
    match = re.search(r'(\w+)=', string)
    result = match.group(1)
    signals.append(result) #'HADDR', 'HBURST'...

coverpoints = list(zip(signals, column_data))

cross_data = []
for row in ws.iter_rows(min_row=11, max_row=16, min_col=1, max_col=2, values_only=True):
    for cell_value in row:
        if cell_value is not None:
            cross_data.append(cell_value)

before_vs = []
after_vs = []
for element in cross_data:
    parts = element.split("_vs_")
    before_vs.append(parts[0])
    after_vs.append(parts[1])

print(before_vs)
print(after_vs)
cross_list = list(zip(before_vs, after_vs))

with open('fifo_cov.sv', 'w') as file:
    file.write('`define cov\n')
    file.write('`ifdef COVERAGE_EN\n\n')
    file.write('covergroup ' + covergroup_name + ' with function sample(fifo_transaction item);\n')
    file.write('option.per_instance = ' + str(1) + ';\n')
    file.write('option.name = "' + covergroup_name + '";\n\n')
    for cp_name, cp_bins in coverpoints:
        file.write('    coverpoint_' + cp_name + '      : coverpoint item.' + cp_name + '\n')
        file.write('    {\n')
        for i, bin_range in enumerate(cp_bins.split('{')[1].split('}')[0].split(',')):
            file.write('        bins    ' + cp_name + '_' + str(i) + ' = {' + bin_range.strip() + '};\n')
        file.write('    }\n')
    for bf, af in cross_list:
        file.write(bf + '_vs_' + af + ' : cross item.' + bf + ', item.' + af + ';\n')
    file.write('endgroup\n')
    file.write('\n\n`endif')

#coverpoints list: ('HRDATA', 'SPLIT_BINS:HRDATA= { [32'h0:32'h66666665], 32'h66666666:32'hcccccccb, 32'hcccccccc:32'hffffffff}'), ...
"""
write_enable_vs_full : cross item.wr_en, item.full;
read_enable_vs_full : cross item.rd_en, item.full;
read_enable_vs_empty : cross item.rd_en, item.empty;
write_enable_vs_empty : cross item.wr_en, item.empty;
write_enable_vs_wrap_on_full : cross item.wr_en, item.wrap_on_full;
read_enable_vs_wrap_on_full : cross item.rd_en, item.wrap_on_full;

"""


#coverpoints list: ('din', 'SPLIT_BINS:din= { [8'h00:8'h7F], [8'h80:8'hBF], [8'hC0:8'hFF]}'), ...
